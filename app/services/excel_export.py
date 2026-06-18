import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.services.pipeline import PipelineResult


# ── Style Constants ──

FONT_BODY = Font(name="Arial", size=10)
FONT_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_GROUP_HEADER = Font(name="Arial", size=10, bold=True, color="1F4E79")
FONT_TITLE = Font(name="Arial", size=14, bold=True, color="1F4E79")
FONT_SUBTITLE = Font(name="Arial", size=10, color="666666")
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_ERROR = Font(name="Arial", size=10, color="CC0000")
FONT_SCORE = Font(name="Arial", size=11, bold=True)

FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_LEAD_INFO = PatternFill("solid", fgColor="D6E4F0")
FILL_LINKEDIN = PatternFill("solid", fgColor="DDE8F0")
FILL_COMPANY = PatternFill("solid", fgColor="E2EFDA")
FILL_ROLE = PatternFill("solid", fgColor="FCE4D6")
FILL_PAIN = PatternFill("solid", fgColor="FFF2CC")
FILL_META = PatternFill("solid", fgColor="F2F2F2")
FILL_ALT_ROW = PatternFill("solid", fgColor="F8F9FA")
FILL_LOW_CONFIDENCE = PatternFill("solid", fgColor="FFF0F0")
FILL_LOW_CONFIDENCE_ALT = PatternFill("solid", fgColor="FFE4E4")
FONT_FLAG = Font(name="Arial", size=10, bold=True, color="CC0000")

MIN_CONFIDENCE = 5
MIN_EXTRACTED_URLS = 2

ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
ALIGN_CENTER = Alignment(horizontal="center", vertical="top")
ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


# ── Column Definitions ──
# (header_label, group_fill, column_width)

COLUMNS = [
    # Lead Info (1-4)
    ("Name", FILL_LEAD_INFO, 20),
    ("Company", FILL_LEAD_INFO, 20),
    ("Email", FILL_LEAD_INFO, 25),
    ("LinkedIn", FILL_LEAD_INFO, 15),
    # LinkedIn Intel (5-8) — populated when Apify LinkedIn scraping is enabled
    ("Headline", FILL_LINKEDIN, 35),
    ("About", FILL_LINKEDIN, 45),
    ("Skills", FILL_LINKEDIN, 35),
    ("LinkedIn Posts", FILL_LINKEDIN, 65),
    # Company Snapshot (9-12)
    ("What They Do", FILL_COMPANY, 45),
    ("Industry", FILL_COMPANY, 25),
    ("Key Offerings", FILL_COMPANY, 40),
    ("Size & Stage", FILL_COMPANY, 25),
    # Prospect Role (13-14)
    ("Likely Priorities", FILL_ROLE, 40),
    ("Key Responsibilities", FILL_ROLE, 40),
    # Pain Signals (15-20)
    ("Pain Signal 1", FILL_PAIN, 40),
    ("Evidence 1", FILL_PAIN, 35),
    ("Pain Signal 2", FILL_PAIN, 40),
    ("Evidence 2", FILL_PAIN, 35),
    ("Pain Signal 3", FILL_PAIN, 40),
    ("Evidence 3", FILL_PAIN, 35),
    # Meta (21-26)
    ("Confidence (1-10)", FILL_META, 16),
    ("Data Gaps", FILL_META, 40),
    ("Source URL 1", FILL_META, 45),
    ("Source URL 2", FILL_META, 45),
    ("Quality Flag", FILL_META, 20),
    ("Error", FILL_META, 30),
]


def _get_quality_flag(lead_result) -> str:
    """Flag leads that need manual review."""
    flags = []
    research = lead_result.research or {}
    confidence = research.get("confidence_score", 0)
    jina = lead_result.jina or {}

    if lead_result.error:
        flags.append("PIPELINE ERROR")
    if confidence < MIN_CONFIDENCE:
        flags.append(f"LOW CONFIDENCE ({confidence}/10)")
    if jina.get("extracted", 0) < MIN_EXTRACTED_URLS:
        flags.append(f"THIN DATA ({jina.get('extracted', 0)} sources)")
    if not research:
        flags.append("NO RESEARCH")

    return " | ".join(flags) if flags else ""


def _needs_review(lead_result) -> bool:
    """Check if a lead should be flagged for review."""
    research = lead_result.research or {}
    confidence = research.get("confidence_score", 0)
    jina = lead_result.jina or {}
    return (
        lead_result.error is not None
        or confidence < MIN_CONFIDENCE
        or jina.get("extracted", 0) < MIN_EXTRACTED_URLS
        or not research
    )


def _extract_linkedin_cells(lead_result) -> list:
    """Extract LinkedIn columns from lead result. Returns 4 cells:
    [headline, about, skills, posts_with_engagement]"""
    li = lead_result.linkedin_data or {}
    profile = li.get("profile", {})
    posts = li.get("posts", [])

    headline = profile.get("headline", "")
    about = profile.get("about", "")

    skills = ", ".join(profile.get("skills", []))

    post_blocks = []
    for i, p in enumerate(posts[:5], 1):
        date = p.get("relative_date", p.get("date", ""))
        reactions = p.get("total_reactions", 0)
        comments = p.get("comments", 0)
        reposts = p.get("reposts", 0)
        text = p.get("text", "")

        header = f"[Post {i} — {date} | {reactions} reactions, {comments} comments, {reposts} reposts]"
        post_blocks.append(f"{header}\n{text}")

    posts_cell = "\n\n———\n\n".join(post_blocks)

    return [headline, about, skills, posts_cell]


def _extract_row(lead_result) -> list:
    """Extract a flat row from a LeadResult for the spreadsheet."""
    lead = lead_result.lead
    research = lead_result.research or {}
    source_urls = lead_result.source_urls or []

    company = research.get("company_snapshot", {})
    role = research.get("prospect_role", {})
    pains = research.get("pain_signals", [])

    # Flatten pain signals (up to 3) — no cell left blank
    pain_cells = []
    for i in range(3):
        if i < len(pains):
            pain_cells.append(pains[i].get("challenge", ""))
            pain_cells.append(pains[i].get("evidence", ""))
        elif not pains:
            pain_cells.append("No public complaints, reviews, or news found for this company")
            pain_cells.append("No evidence to cite — no complaints, reviews, or news found")
        else:
            pain_cells.append("No additional pain signals identified")
            pain_cells.append("No further evidence found")

    data_gaps = research.get("data_gaps", [])
    gaps_text = "\n".join(f"• {g}" for g in data_gaps)

    return [
        # Lead Info
        lead.get("name", ""),
        lead.get("company", ""),
        lead.get("email", ""),
        lead.get("linkedin", ""),
        # LinkedIn Intel
        *_extract_linkedin_cells(lead_result),
        # Company Snapshot
        company.get("what_they_do", ""),
        company.get("industry", ""),
        "\n".join(f"• {o}" for o in company.get("key_offerings", [])),
        company.get("size_and_stage", ""),
        # Prospect Role
        role.get("likely_priorities", ""),
        role.get("key_responsibilities", ""),
        # Pain Signals (flattened, may have empty cells if <3 signals)
        *pain_cells,
        # Meta
        research.get("confidence_score", ""),
        gaps_text,
        source_urls[0] if len(source_urls) > 0 else "",
        source_urls[1] if len(source_urls) > 1 else "",
        _get_quality_flag(lead_result),
        lead_result.error or "",
    ]


def export_pipeline_results(pipeline_result: PipelineResult) -> io.BytesIO:
    """Generate a well-structured Excel workbook from pipeline results."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Lead Research"

    # ── Group Header Row (Row 1) ──
    group_spans = [
        ("Lead Info", 1, 4, FILL_LEAD_INFO),
        ("LinkedIn Intel", 5, 8, FILL_LINKEDIN),
        ("Company Snapshot", 9, 12, FILL_COMPANY),
        ("Prospect Role", 13, 14, FILL_ROLE),
        ("Pain Signals", 15, 20, FILL_PAIN),
        ("Meta", 21, 26, FILL_META),
    ]

    for label, start_col, end_col, fill in group_spans:
        if start_col == end_col:
            cell = ws.cell(row=1, column=start_col, value=label)
        else:
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=end_col,
            )
            cell = ws.cell(row=1, column=start_col, value=label)
        cell.font = FONT_GROUP_HEADER
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Column Header Row (Row 2) ──
    for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_HEADER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Data Rows (starting Row 3) ──
    for row_offset, lead_result in enumerate(pipeline_result.results):
        row_num = 3 + row_offset
        row_data = _extract_row(lead_result)
        use_alt = row_offset % 2 == 1
        flagged = _needs_review(lead_result)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_WRAP
            cell.border = THIN_BORDER

            # Row-level color: flagged leads get red tint, normal leads get alternating
            if flagged:
                cell.fill = FILL_LOW_CONFIDENCE_ALT if use_alt else FILL_LOW_CONFIDENCE
            elif use_alt:
                cell.fill = FILL_ALT_ROW

            # Confidence score color coding (col 21)
            if col_idx == 21 and isinstance(value, int):
                cell.alignment = ALIGN_CENTER
                cell.font = FONT_SCORE
                if value >= 8:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                elif value >= 5:
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")
                else:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")

            # Source URLs as clickable hyperlinks (cols 23-24)
            if col_idx in (23, 24) and value:
                cell.hyperlink = value
                cell.font = FONT_LINK

            # LinkedIn as clickable hyperlink (col 4)
            if col_idx == 4 and value:
                cell.hyperlink = value
                cell.value = "Profile"
                cell.font = FONT_LINK

            # Quality flag column bold red (col 25)
            if col_idx == 25 and value:
                cell.font = FONT_FLAG

            # Error column red (col 26)
            if col_idx == 26 and value:
                cell.font = FONT_ERROR

    # ── Freeze Panes ──
    ws.freeze_panes = "E3"

    # ── Row height for data rows ──
    for row_num in range(3, 3 + len(pipeline_result.results)):
        ws.row_dimensions[row_num].height = 150

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 30

    # ── Auto-filter ──
    last_col = get_column_letter(len(COLUMNS))
    last_row = 2 + len(pipeline_result.results)
    ws.auto_filter.ref = f"A2:{last_col}{last_row}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
